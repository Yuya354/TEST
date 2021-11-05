"""
Created on Sun Jan 24 22:14:53 2021
@author: Yuya
"""
import numpy as np
from matplotlib import pyplot as plt
import openpyxl
import math
# ガウス関数を定義
def gauss(x1, amp = 1, loc = 0, sigma = 1):
    return amp * np.exp(-(x1 - loc)**2 / (2*sigma**2))

# フォントの種類とサイズを設定する。
plt.rcParams['font.size'] = 14
plt.rcParams['font.family'] = 'Times New Roman'
 
# 目盛を内側にする。
plt.rcParams['xtick.direction'] = 'in'
plt.rcParams['ytick.direction'] = 'in'
 
fig = plt.figure()
ax1 = plt.subplot(111)
 
# グラフの上下左右に目盛線を付ける。
ax1.yaxis.set_ticks_position('both')
ax1.xaxis.set_ticks_position('both')
 
# 軸のラベルを設定する。
ax1.set_ylabel('Intensity')
ax1.set_xlabel('Band Energy') 

# データの範囲と刻み目盛を明示する。
ax1.set_xticks(np.arange(94, 108, 1))
ax1.set_xlim(93, 109)
 
ax1.invert_xaxis()

# ブックを取得
WB = openpyxl.load_workbook('data.xlsx')
# シートを取得
WS = WB['Sheet1']

for i in range(0,10000):
    #平均値(loc),標準偏差(scale)
    x1 = np.arange(0,120,0.05)
    y_1 = np.random.normal(loc=99.25, scale=1, size = len(x1))
    z_1 = np.random.normal(loc=0.2545, scale=((0.02545*3)/3) ,size = len(x1))
    y1 = gauss(x1,amp = 1, loc = y_1[1200], sigma = z_1[1200])

    y_2 = np.random.normal(loc=y_1[1200]+0.7, scale=0.01667, size = len(x1))
    y2 = gauss(x1,amp = 0.5, loc = y_2, sigma = z_1[1200])

    y_3 = np.random.normal(loc=y_1[1200]+4.5 , scale=0.1667, size = len(x1))
    rand = np.random.normal(loc = np.random.uniform(1*0.01,1*2), scale = 0.01667, size = len(x1))
    z_2 = np.random.uniform(2.0*z_1,2.5*z_1)
    y3 = gauss(x1,amp = rand, loc = y_3[1200], sigma = z_2[1200])

    y4 = (y1 + y2 + y3)*5000
    y_4 = y4[1910:2160]
    
    #FWHM = 2*sgm*√(2*ln2)
    High1 = 5000
    High2 = High1*0.5
    High3 = rand[1200]*5000
    FWHM1 = 2*z_1[1200]*np.sqrt(2*math.log(2))
    FWHM2 = 2*z_1[1200]*np.sqrt(2*math.log(2))
    FWHM3 = 2*z_2[1200]*np.sqrt(2*math.log(2))

    list = []
    list.append(y_1[1200])
    list.append(y_2[1200])
    list.append(y_3[1200])
    list.append(FWHM1)
    list.append(FWHM2)
    list.append(FWHM3)
    list.append(High1)
    list.append(High2)
    list.append(High3)
    
#シートへの書き込み　rowに対象セルの行、columnに対象セルの列を指定
    for j in range(0,250):
        WS.cell(row=i+2,column=j+1).value = (y_4[j])
        
    for m in range(0,9):
        WS.cell(row=i+2,column=m+251).value = (list[m])       
# 保存する
WB.save('data.xlsx')

plt.plot(x1,y4,label='PDF', lw=1,color="black")
plt.show()


